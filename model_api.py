import requests
import streamlit as st
import openai
from openai import OpenAI
import openpyxl
from openpyxl import Workbook
from datetime import datetime
from GSAR_functions import functions

class UniversalModelInterface:

    def __init__(self):
        # Try to get API key from secrets if not set in openai module
        api_key = openai.api_key
        try:
            if not api_key and 'openai' in st.secrets:
                api_key = st.secrets["openai"]["api_key"]
        except:
            pass
        self.client = OpenAI(api_key=api_key)

    def initialize_excel(self, file_path):
        try:
            wb = openpyxl.load_workbook(file_path)
        except FileNotFoundError:
            wb = Workbook()
            ws = wb.active
            ws.append(['Question', 'Response', 'Date', 'Count'])
            wb.save(file_path)
        return wb
    
    def update_excel(self, wb, file_path, question, response):
        ws = wb.active
        current_date = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        question_found = False
        row_index = None

        for r_index, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):  # Skip header row
            if row[0] == question:
                question_found = True
                row_index = r_index
                break
        
        if question_found:
            ws.cell(row=row_index, column=3, value=current_date)  # Update date
            count = ws.cell(row=row_index, column=4).value or 0
            ws.cell(row=row_index, column=4, value=count + 1)  # Update count
        else:
            ws.append([question, response, current_date, 1])  # New question

        wb.save(file_path)

    def get_response(self, model_name, prompt_question, file_path, prompt):
        # Make sure the file exists
        import os
        os.makedirs(os.path.dirname(file_path), exist_ok=True)
        
        wb = self.initialize_excel(file_path)

        if model_name == "llama2":
            res = self._local_server_request(prompt_question)
            self.update_excel(wb, file_path, prompt, res)
            return res
        elif model_name == "gpt-4":
            res = self._openai_request_content(prompt_question)
            self.update_excel(wb, file_path, prompt, res)
            return res
        elif model_name == "gpt-4-response":
            res = self._openai_request_response(prompt_question)
            response_message = res.choices[0].message
            if hasattr(response_message, "function_call") and response_message.function_call:
                self.update_excel(wb, file_path, prompt, "Document Created")
            else:
                formatted_response = response_message.content
                self.update_excel(wb, file_path, prompt, formatted_response)
            return res
        else:
            raise ValueError(f"Unsupported model: {model_name}")


    def _local_server_request(self, prompt_question):
        url = "https://jjptfo2swt2i4p-8080.proxy.runpod.net/generate"
        headers = {
            "Content-Type": "application/json"
        }
        data = {
            "inputs": prompt_question,
            "parameters": {
                "max_new_tokens": 20
            }
        }

        response = requests.post(url, headers=headers, json=data)
        response.raise_for_status()  # This will raise an exception for HTTP error responses
        print(response.json())
        return response.json()['generated_text'].strip()

    def _openai_request_content(self, prompt_question):
        response = self.client.chat.completions.create(
            model='gpt-4',
            messages=[
                {"role": "user", "content": prompt_question}
            ],
            temperature=0.0
        )
        return response.choices[0].message.content
    
    def _openai_request_response(self, prompt_question):
        response = self.client.chat.completions.create(
            model='gpt-4',
            messages=[
                {"role": "user", "content": prompt_question}
            ],
            functions = functions.functions_call,
            temperature=0.0
        )
        return response